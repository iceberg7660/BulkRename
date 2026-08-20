import os
import openpyxl
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout,
                              QHBoxLayout, QLabel, QLineEdit, QFileDialog, QComboBox,
                              QMessageBox, QMenuBar, QAction, QWidget, QCheckBox)

class RenameApp(QMainWindow):
    MAX_CUSTOM_FIELDS = 5

    def __init__(self):
        super().__init__()
        # ── 数据模型 ──
        self.roster_has_headers = False
        self.roster_columns = []       # 列名列表，如 ['学号', '姓名'] 或 ['列1', '列2']
        self.roster_match_column = 0   # 匹配依据列索引
        self.roster_rows = []          # [{'学号': 2021001, '姓名': '张三'}, ...]
        self.custom_field_widgets = [] # [{'name_edit': QLE, 'value_edit': QLE, 'row': QW, 'remove_btn': QPB}]
        self.filename_combos = []      # 四个文件名组成下拉框

        self.initUI()

    # ═══════════════════════════════════════════════════════════════
    # UI 初始化
    # ═══════════════════════════════════════════════════════════════

    def initUI(self):
        self.setWindowTitle('批量重命名工具')
        self.setGeometry(100, 100, 620, 650)

        # ── 菜单栏 ──
        menubar = QMenuBar(self)
        self.setMenuBar(menubar)
        about_action = QAction('简介', self)
        about_action.triggered.connect(self.show_about_dialog)
        menubar.addAction(about_action)
        qt_about_action = QAction('Qt5介绍', self)
        qt_about_action.triggered.connect(self.show_qt_about_dialog)
        menubar.addAction(qt_about_action)

        # ── 中心部件 + 主布局 ──
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # ── Step 1：名单文件 ──
        self.label_roster = QLabel('Step1：请选择名单文件')
        layout.addWidget(self.label_roster)

        roster_btn_layout = QHBoxLayout()
        self.btn_roster = QPushButton('选择名单文件')
        self.btn_roster.clicked.connect(self.select_roster)
        self.btn_preset_roster = QPushButton('使用预设名单')
        self.btn_preset_roster.clicked.connect(self.use_preset_roster)
        roster_btn_layout.addWidget(self.btn_roster)
        roster_btn_layout.addWidget(self.btn_preset_roster)
        layout.addLayout(roster_btn_layout)

        self.line_roster = QLineEdit()
        layout.addWidget(self.line_roster)

        # 表头复选框（独立一行）
        self.cb_has_headers = QCheckBox('第一行为表头')
        self.cb_has_headers.setEnabled(False)
        self.cb_has_headers.stateChanged.connect(self._on_headers_checkbox_changed)
        layout.addWidget(self.cb_has_headers)

        # 匹配依据列（独立一行，突出显示）
        match_row = QHBoxLayout()
        match_label = QLabel('匹配依据列：')
        match_label.setStyleSheet('font-weight: bold;')
        match_row.addWidget(match_label)
        self.combo_match_column = QComboBox()
        self.combo_match_column.setEnabled(False)
        self.combo_match_column.setMinimumWidth(200)
        self.combo_match_column.currentIndexChanged.connect(self._on_match_column_changed)
        match_row.addWidget(self.combo_match_column)
        match_row.addStretch()
        layout.addLayout(match_row)

        # ── Step 2：文件夹 ──
        self.label_folder = QLabel('Step2：选择需要重命名文件的文件夹*')
        layout.addWidget(self.label_folder)
        self.btn_folder = QPushButton('选择文件夹')
        self.btn_folder.clicked.connect(self.select_folder)
        layout.addWidget(self.btn_folder)
        self.line_folder = QLineEdit()
        layout.addWidget(self.line_folder)

        # ── Step 3：新文件夹名 ──
        self.label_new_folder = QLabel('Step3：新的文件夹名字（可选）')
        layout.addWidget(self.label_new_folder)
        self.line_new_folder = QLineEdit()
        layout.addWidget(self.line_new_folder)

        # ── Step 4：自定义字段（动态） ──
        self.label_custom_fields = QLabel('Step4：自定义字段（可选，最多5个）')
        layout.addWidget(self.label_custom_fields)

        self.btn_add_custom = QPushButton('＋ 添加自定义字段')
        self.btn_add_custom.clicked.connect(self._add_custom_field_row)
        layout.addWidget(self.btn_add_custom)

        self.custom_fields_container = QWidget()
        self.custom_fields_layout = QVBoxLayout(self.custom_fields_container)
        self.custom_fields_layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.custom_fields_container)

        # ── Step 5：文件名组成（动态） ──
        self.label_final_name = QLabel('Step5：最终文件名组成')
        layout.addWidget(self.label_final_name)

        # 动态组合框容器
        self.filename_combos_container = QWidget()
        self.filename_combos_layout = QHBoxLayout(self.filename_combos_container)
        self.filename_combos_layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.filename_combos_container)

        # ── 文件名预览 ──
        self.label_preview = QLabel('最终文件名预览')
        layout.addWidget(self.label_preview)
        self.line_preview = QLineEdit()
        self.line_preview.setReadOnly(True)
        layout.addWidget(self.line_preview)

        # ── 重命名按钮 ──
        self.btn_rename = QPushButton('重命名')
        self.btn_rename.clicked.connect(self.rename_files)
        layout.addWidget(self.btn_rename)

        # ── 初始化：1个组合框 ──
        combo = QComboBox()
        combo.currentIndexChanged.connect(self.update_preview)
        self.filename_combos.append(combo)
        self.filename_combos_layout.addWidget(combo)
        self.rebuild_combo_options()

    # ═══════════════════════════════════════════════════════════════
    # 对话框
    # ═══════════════════════════════════════════════════════════════

    def show_about_dialog(self):
        about_text = (
            "使用方法：\n"
            "1. 选择名单文件（Excel格式，支持 .xlsx / .xls）。\n"
            "2. 程序自动识别第一行是否为表头，可通过复选框纠正。\n"
            "3. 表头列名或自动生成的列名将作为重命名时可选的字段。\n"
            "4. 选择需要重命名文件的文件夹。\n"
            "5. 可选：输入新的文件夹名字。\n"
            "6. 可选：点击「＋」添加自定义字段（最多5个），\n"
            "   每个字段包含名称和值。\n"
            "7. 选择匹配依据列（用于在文件名中定位原文件）。\n"
            "8. 选择最终文件名的组成顺序。\n"
            "9. 点击「重命名」按钮开始重命名。\n\n"
            "* 本程序使用vibecoding，使用 Python + PyQt5 开发。\n"
            "* 本程序不会更改文件的格式和内容，保留原文件后缀名。\n"
            "* 匹配原理：将名单匹配依据列的值与文件名字符串比对，\n"
            "  文件名中包含该值即视为匹配。\n"
            "* 匹配依据列若有重复值，仅最先出现的一行参与重命名，\n"
            "  其余重复行跳过，重命名结束后弹窗给出详细说明。\n"
            "* 本程序仅供学习交流使用，不得用于商业用途。\n"
        )
        QMessageBox.about(self, '简介', about_text)

    def show_qt_about_dialog(self):
        QMessageBox.aboutQt(self, '关于Qt')

    # ═══════════════════════════════════════════════════════════════
    # 名单加载与解析
    # ═══════════════════════════════════════════════════════════════

    def select_roster(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择名单文件', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            self.line_roster.setText(file_path)
            self.load_roster(file_path)

    def use_preset_roster(self):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_dir, '名单.xlsx')
        if os.path.exists(file_path):
            self.line_roster.setText(file_path)
            self.load_roster(file_path)
        else:
            QMessageBox.warning(self, '警告', '未找到预设名单文件（名单.xlsx）')

    def load_roster(self, file_path, force_has_headers=None):
        """解析 Excel 名单文件，填充数据模型。"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            if sheet is None:
                self._set_roster_error('无法读取工作表')
                return

            if sheet.max_row < 1 or sheet.max_column < 1:
                self._set_roster_error('文件为空或格式不正确')
                return

            # ── 判断是否有表头 ──
            if force_has_headers is not None:
                self.roster_has_headers = force_has_headers
            else:
                self.roster_has_headers = self._detect_has_headers(sheet)

            # 更新复选框（不触发信号）
            self.cb_has_headers.blockSignals(True)
            self.cb_has_headers.setChecked(self.roster_has_headers)
            self.cb_has_headers.blockSignals(False)
            self.cb_has_headers.setEnabled(True)

            # ── 构建列名 ──
            if self.roster_has_headers:
                self.roster_columns = []
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=1, column=c).value
                    self.roster_columns.append(str(val) if val is not None else f'列{c}')
                data_start_row = 2
            else:
                self.roster_columns = [f'列{c}' for c in range(1, sheet.max_column + 1)]
                data_start_row = 1

            # ── 构建数据行 ──
            self.roster_rows = []
            for r in range(data_start_row, sheet.max_row + 1):
                row_dict = {}
                for c in range(1, sheet.max_column + 1):
                    col_name = self.roster_columns[c - 1]
                    val = sheet.cell(row=r, column=c).value
                    row_dict[col_name] = val
                # 跳过全空行
                if any(v is not None for v in row_dict.values()):
                    row_dict['__excel_row__'] = r  # 内部键：记录 Excel 行号，用于结果报告
                    self.roster_rows.append(row_dict)

            if not self.roster_rows:
                self._set_roster_error('名单文件中没有数据行')
                return

            # ── 自动选择匹配列 ──
            self._auto_select_match_column()

            # ── 填充匹配列下拉框 ──
            self.combo_match_column.blockSignals(True)
            self.combo_match_column.clear()
            self.combo_match_column.addItems(self.roster_columns)
            self.combo_match_column.setCurrentIndex(self.roster_match_column)
            self.combo_match_column.blockSignals(False)
            self.combo_match_column.setEnabled(True)

            # ── 刷新文件名组成下拉框 ──
            self._sync_combo_count_to_fields()

            # ── 检测匹配依据列是否有重复值（有则弹窗提醒）──
            self._check_match_column_duplicates()

            # ── 状态提示 ──
            header_status = '有表头' if self.roster_has_headers else '无表头'
            self.label_roster.setText(
                f'Step1：请选择名单文件 — 识别成功（{header_status}，'
                f'{len(self.roster_columns)}列，{len(self.roster_rows)}行）')
            self.label_roster.setStyleSheet('color: green;')

        except Exception as e:
            self._set_roster_error(f'读取错误: {str(e)}')

    def _set_roster_error(self, msg):
        self.roster_columns = []
        self.roster_rows = []
        self.roster_has_headers = False
        self.cb_has_headers.setEnabled(False)
        self.combo_match_column.clear()
        self.combo_match_column.setEnabled(False)
        self.label_roster.setText(f'Step1：请选择名单文件 — {msg}')
        self.label_roster.setStyleSheet('color: red;')

    def _detect_has_headers(self, sheet):
        """试探性地判断第一行是否为表头。"""
        if sheet.max_row < 2:
            return False

        row1_values = []
        for c in range(1, sheet.max_column + 1):
            row1_values.append(sheet.cell(row=1, column=c).value)

        row2_values = []
        for c in range(1, sheet.max_column + 1):
            row2_values.append(sheet.cell(row=2, column=c).value)

        # 第1行：所有非空单元格都不是纯数字
        def is_numeric(v):
            if v is None:
                return False
            if isinstance(v, (int, float)):
                return True
            if isinstance(v, str):
                stripped = v.strip().replace('.', '').replace('-', '')
                return stripped.isdigit()
            return False

        row1_all_text = all(
            v is None or not is_numeric(v) for v in row1_values
        )
        # 第2行：至少有一个数字单元格
        row2_has_numeric = any(is_numeric(v) for v in row2_values)

        return row1_all_text and row2_has_numeric

    def _auto_select_match_column(self):
        """选择包含中文字符比例最高的列作为默认匹配列。"""
        best_col = 0
        best_ratio = -1

        for ci, col_name in enumerate(self.roster_columns):
            chinese_count = 0
            total = 0
            for row in self.roster_rows:
                val = row.get(col_name)
                if val is not None:
                    total += 1
                    s = str(val)
                    if any('一' <= ch <= '鿿' for ch in s):
                        chinese_count += 1
            ratio = chinese_count / total if total > 0 else 0
            if ratio > best_ratio:
                best_ratio = ratio
                best_col = ci

        self.roster_match_column = best_col

    # ═══════════════════════════════════════════════════════════════
    # 表头 / 匹配列 回调
    # ═══════════════════════════════════════════════════════════════

    def _on_headers_checkbox_changed(self, state):
        file_path = self.line_roster.text()
        if not file_path:
            return
        reply = QMessageBox.question(
            self, '确认',
            '更改表头设置将重新解析名单，文件名组成选择可能被重置。是否继续？',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.load_roster(file_path, force_has_headers=(state == 2))  # Qt.Checked

    def _on_match_column_changed(self, index):
        if index >= 0:
            self.roster_match_column = index
            # 切换匹配依据列后立即检测该列是否有重复值
            self._check_match_column_duplicates()

    # ═══════════════════════════════════════════════════════════════
    # 匹配依据列重复检测
    # ═══════════════════════════════════════════════════════════════

    def _normalize_match_value(self, raw):
        """归一化匹配依据列的单元格值。

        与重命名匹配逻辑保持一致：None → 空串；整数值的 float → 整数字符串
        （避免 Excel 数字列的 2021115006.0 与字符串 2021115006 被判为不同值）。
        """
        if raw is None:
            return ''
        if isinstance(raw, float) and raw == int(raw):
            return str(int(raw))
        return str(raw)

    def _find_duplicate_match_values(self):
        """检测当前匹配依据列中的重复值。

        Returns:
            dict: {匹配值: [在 roster_rows 中的行索引列表]}，
            仅包含出现次数大于 1 的值，按首次出现顺序排列。
        """
        if not self.roster_columns or not self.roster_rows:
            return {}
        if self.roster_match_column >= len(self.roster_columns):
            return {}
        match_col_name = self.roster_columns[self.roster_match_column]

        groups = {}
        order = []
        for idx, row_data in enumerate(self.roster_rows):
            val = self._normalize_match_value(row_data.get(match_col_name))
            if not val:
                continue  # 空值不参与重复检测（它们本来就匹配不到文件）
            if val not in groups:
                groups[val] = []
                order.append(val)
            groups[val].append(idx)

        return {v: groups[v] for v in order if len(groups[v]) > 1}

    def _check_match_column_duplicates(self):
        """导入名单 / 切换匹配依据列后，检测该列重复值并弹窗提醒。"""
        duplicates = self._find_duplicate_match_values()
        if not duplicates:
            return

        match_col_name = self.roster_columns[self.roster_match_column]
        lines = []
        items = list(duplicates.items())
        for i, (val, row_indices) in enumerate(items):
            if i >= 10:
                lines.append(f'... 及其他 {len(items) - 10} 个重复值')
                break
            excel_rows = '、'.join(
                str(self.roster_rows[ri].get('__excel_row__', '?')) for ri in row_indices)
            lines.append(f'「{val}」出现 {len(row_indices)} 次（Excel 第 {excel_rows} 行）')

        msg = (
            f'检测到匹配依据列「{match_col_name}」中存在重复值：\n\n'
            + '\n'.join(lines)
            + '\n\n默认处理规则：每个重复值仅名单中最先出现的一行参与重命名，'
            '其余重复行将被跳过（不做处理）。\n'
            '重命名结束后的结果弹窗中会给出详细说明。'
        )
        QMessageBox.warning(self, '匹配依据列存在重复', msg)

    def _describe_row(self, row_data):
        """生成名单行的可读描述（用于结果报告，排除内部键）。"""
        excel_row = row_data.get('__excel_row__')
        parts = []
        for k, v in row_data.items():
            if k == '__excel_row__':
                continue
            s = self._normalize_match_value(v)
            parts.append(f'{k}={s if s else "（空）"}')
        desc = '，'.join(parts)
        if excel_row is not None:
            return f'Excel 第 {excel_row} 行（{desc}）'
        return desc

    # ═══════════════════════════════════════════════════════════════
    # 文件夹选择
    # ═══════════════════════════════════════════════════════════════

    def select_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, '选择文件夹')
        if folder_path:
            self.line_folder.setText(folder_path)

    # ═══════════════════════════════════════════════════════════════
    # 自定义字段（动态）
    # ═══════════════════════════════════════════════════════════════

    def _add_custom_field_row(self):
        if len(self.custom_field_widgets) >= self.MAX_CUSTOM_FIELDS:
            QMessageBox.warning(self, '提示', f'最多只能添加 {self.MAX_CUSTOM_FIELDS} 个自定义字段。')
            return

        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 2, 0, 2)

        name_edit = QLineEdit()
        name_edit.setPlaceholderText('字段名（如：作业名）')
        name_edit.setMinimumWidth(120)
        name_edit.textChanged.connect(lambda: self.rebuild_combo_options())
        name_edit.editingFinished.connect(lambda ed=name_edit: self._validate_custom_field_name(ed))
        row_layout.addWidget(name_edit)

        value_edit = QLineEdit()
        value_edit.setPlaceholderText('字段值（如：作业一）')
        value_edit.textChanged.connect(self.update_preview)
        row_layout.addWidget(value_edit)

        remove_btn = QPushButton('－')
        remove_btn.setFixedWidth(30)
        row_layout.addWidget(remove_btn)

        entry = {
            'name_edit': name_edit,
            'value_edit': value_edit,
            'row_widget': row_widget,
            'remove_btn': remove_btn,
        }
        self.custom_field_widgets.append(entry)

        # 绑定删除回调
        idx = len(self.custom_field_widgets) - 1
        remove_btn.clicked.connect(lambda: self._remove_custom_field_row(idx))

        self.custom_fields_layout.addWidget(row_widget)

        # 更新按钮状态
        self._update_add_button_state()
        # 刷新组合框（数量跟随字段变化）
        self._sync_combo_count_to_fields()

    def _validate_custom_field_name(self, name_edit):
        """检查自定义字段名是否与名单列或其他自定义字段重复。"""
        name = name_edit.text().strip()
        if not name:
            return  # 空名字不检查

        # 找到当前字段的索引（用于排除自身）
        current_idx = None
        for i, entry in enumerate(self.custom_field_widgets):
            if entry['name_edit'] is name_edit:
                current_idx = i
                break

        # 检查是否与名单列名重复
        if name in self.roster_columns:
            QMessageBox.warning(self, '字段名重复',
                                f'字段名「{name}」与名单中的列名重复，请使用不同的名称。')
            name_edit.clear()
            name_edit.setFocus()
            return

        # 检查是否与其他自定义字段名重复
        for i, entry in enumerate(self.custom_field_widgets):
            if i != current_idx:
                other_name = entry['name_edit'].text().strip()
                if other_name and other_name == name:
                    QMessageBox.warning(self, '字段名重复',
                                        f'字段名「{name}」已存在，请使用不同的名称。')
                    name_edit.clear()
                    name_edit.setFocus()
                    return

    def _remove_custom_field_row(self, index):
        if index < 0 or index >= len(self.custom_field_widgets):
            return
        entry = self.custom_field_widgets.pop(index)
        entry['row_widget'].setParent(None)

        # 重新绑定所有删除按钮的索引
        for i, ent in enumerate(self.custom_field_widgets):
            ent['remove_btn'].clicked.disconnect()
            ent['remove_btn'].clicked.connect(
                lambda _, idx=i: self._remove_custom_field_row(idx))

        self._update_add_button_state()
        self._sync_combo_count_to_fields()

    def _update_add_button_state(self):
        if len(self.custom_field_widgets) >= self.MAX_CUSTOM_FIELDS:
            self.btn_add_custom.setEnabled(False)
            self.btn_add_custom.setText('自定义字段已满')
        else:
            self.btn_add_custom.setEnabled(True)
            self.btn_add_custom.setText('＋ 添加自定义字段')

    # ═══════════════════════════════════════════════════════════════
    # 文件名组成下拉框
    # ═══════════════════════════════════════════════════════════════

    def _sync_combo_count_to_fields(self):
        """根据列名和自定义字段总数，自动调整组合框数量。"""
        total_fields = len(self.roster_columns) + len(self.custom_field_widgets)
        target = max(1, total_fields)

        while len(self.filename_combos) < target:
            combo = QComboBox()
            combo.currentIndexChanged.connect(self.update_preview)
            self.filename_combos.append(combo)
            self.filename_combos_layout.addWidget(combo)

        while len(self.filename_combos) > target:
            combo = self.filename_combos.pop()
            combo.setParent(None)

        self.rebuild_combo_options()

    def rebuild_combo_options(self):
        """根据当前列名和自定义字段，刷新所有文件名组成下拉框的选项。"""
        options = list(self.roster_columns)
        for entry in self.custom_field_widgets:
            name = entry['name_edit'].text().strip()
            if name:
                options.append(name)
        options.append('（空）')

        for combo in self.filename_combos:
            old_text = combo.currentText()
            old_index = combo.currentIndex()
            combo.blockSignals(True)
            combo.clear()
            combo.addItems(options)
            # 优先按名称恢复，名称变化时按索引恢复
            idx = combo.findText(old_text)
            if idx >= 0:
                combo.setCurrentIndex(idx)
            elif old_index < len(options):
                combo.setCurrentIndex(old_index)
            combo.blockSignals(False)

        self.update_preview()

    # ═══════════════════════════════════════════════════════════════
    # 字段值解析（消除重复的 if-elif）
    # ═══════════════════════════════════════════════════════════════

    def _resolve_field_value(self, field_name, row_data=None):
        """将下拉框中选中的字段名映射为实际值。

        Args:
            field_name: 下拉框当前文本
            row_data:  当前行的 dict（预览模式下为 None）

        Returns:
            解析后的字符串
        """
        if field_name == '（空）':
            return ''

        # 检查自定义字段
        for entry in self.custom_field_widgets:
            if entry['name_edit'].text().strip() == field_name:
                return entry['value_edit'].text()

        # 检查名单列
        if row_data is not None and field_name in row_data:
            val = row_data[field_name]
            if val is None:
                return ''
            if isinstance(val, float) and val == int(val):
                return str(int(val))
            return str(val)

        # 预览模式下的占位显示
        return field_name

    # ═══════════════════════════════════════════════════════════════
    # 工具方法
    # ═══════════════════════════════════════════════════════════════

    @staticmethod
    def _sanitize_filename(name):
        """移除文件名中的非法字符（Windows）。"""
        illegal = r'\/:*?"<>|'
        for ch in illegal:
            name = name.replace(ch, '_')
        return name

    # ═══════════════════════════════════════════════════════════════
    # 预览
    # ═══════════════════════════════════════════════════════════════

    def update_preview(self):
        parts = []
        for combo in self.filename_combos:
            text = combo.currentText()
            parts.append(self._resolve_field_value(text, row_data=None))
        self.line_preview.setText(''.join(parts))

    # ═══════════════════════════════════════════════════════════════
    # 重命名
    # ═══════════════════════════════════════════════════════════════

    def rename_files(self):
        roster_path = self.line_roster.text()
        folder_path = self.line_folder.text()
        new_folder_name = self.line_new_folder.text()

        if not roster_path or not folder_path:
            QMessageBox.warning(self, '警告', '请选择名单文件和文件夹路径')
            return

        if not self.roster_rows or not self.roster_columns:
            QMessageBox.warning(self, '警告', '名单数据为空，请先正确加载名单文件')
            return

        # ── 重命名前检查自定义字段名是否重复 ──
        seen_names = set()
        for entry in self.custom_field_widgets:
            name = entry['name_edit'].text().strip()
            if name in seen_names:
                QMessageBox.warning(self, '字段名重复',
                                    f'自定义字段名「{name}」重复，请修改后再重命名。')
                entry['name_edit'].setFocus()
                return
            if name and name in self.roster_columns:
                QMessageBox.warning(self, '字段名冲突',
                                    f'自定义字段名「{name}」与名单列名重复，请修改后再重命名。')
                entry['name_edit'].setFocus()
                return
            if name:
                seen_names.add(name)

        try:
            filenames = os.listdir(folder_path)
            if not filenames:
                QMessageBox.warning(self, '警告', '文件夹中没有文件')
                return

            match_col_name = self.roster_columns[self.roster_match_column] \
                if self.roster_match_column < len(self.roster_columns) else self.roster_columns[0]

            # ── 匹配值重复预检：每个重复值仅第一行参与重命名，其余跳过 ──
            duplicate_groups = self._find_duplicate_match_values()
            skipped_dup_indices = set()
            for row_indices in duplicate_groups.values():
                skipped_dup_indices.update(row_indices[1:])

            row_outcomes = {}  # roster_rows 索引 → 该行处理结果（success/error/unmatched/skipped_dup）

            for row_idx, row_data in enumerate(self.roster_rows):
                if row_idx in skipped_dup_indices:
                    row_outcomes[row_idx] = {'status': 'skipped_dup'}
                    continue

                match_value = self._normalize_match_value(row_data.get(match_col_name))
                if not match_value:
                    row_outcomes[row_idx] = {'status': 'unmatched'}
                    continue

                matched = False
                for name_in_list in filenames:
                    if match_value in name_in_list:
                        ori_path = os.path.join(folder_path, name_in_list)
                        if not os.path.exists(ori_path):
                            continue  # 已被前面的行重命名掉
                        ext = os.path.splitext(ori_path)[-1]

                        parts = []
                        for combo in self.filename_combos:
                            field_name = combo.currentText()
                            parts.append(self._resolve_field_value(field_name, row_data=row_data))

                        new_name = self._sanitize_filename(''.join(parts)) + ext
                        new_path = os.path.join(folder_path, new_name)

                        try:
                            os.rename(ori_path, new_path)
                            row_outcomes[row_idx] = {
                                'status': 'success', 'ori': name_in_list, 'new': new_name}
                        except Exception as e:
                            # 不再逐个弹窗，统一汇总到结束后的结果报告
                            row_outcomes[row_idx] = {
                                'status': 'error', 'ori': name_in_list, 'error': str(e)}
                        matched = True
                        break

                if not matched:
                    row_outcomes[row_idx] = {'status': 'unmatched'}

            # ── 重命名文件夹 ──
            folder_note = ''
            if new_folder_name:
                new_folder_path = os.path.join(os.path.dirname(folder_path), new_folder_name)
                try:
                    os.rename(folder_path, new_folder_path)
                    folder_path = new_folder_path
                except FileExistsError:
                    folder_note = f'⚠ 文件夹未能重命名：目标文件夹「{new_folder_name}」已存在。'
                except FileNotFoundError:
                    folder_note = '⚠ 文件夹未能重命名：文件夹路径无效。'

            # ── 结果报告：成功数 / 失败数 / 失败原因 / 重复值处理详情 ──
            self._show_result_report(match_col_name, duplicate_groups, row_outcomes, folder_note)

            if folder_path and os.path.isdir(folder_path):
                os.startfile(folder_path)

        except FileNotFoundError:
            QMessageBox.warning(self, '警告', '名单文件路径无效')
        except Exception as e:
            QMessageBox.warning(self, '警告', f'发生错误: {str(e)}')

    def _show_result_report(self, match_col_name, duplicate_groups, row_outcomes, folder_note=''):
        """重命名结束后的汇总弹窗：成功数、失败数、失败原因、重复值处理详情。"""
        success_items = [(ri, o) for ri, o in row_outcomes.items() if o['status'] == 'success']
        error_items = [(ri, o) for ri, o in row_outcomes.items() if o['status'] == 'error']
        unmatched_items = [(ri, o) for ri, o in row_outcomes.items() if o['status'] == 'unmatched']
        skipped_count = sum(1 for o in row_outcomes.values() if o['status'] == 'skipped_dup')
        failed_count = len(error_items) + len(unmatched_items) + skipped_count

        lines = [
            f'名单共 {len(self.roster_rows)} 行',
            f'成功重命名：{len(success_items)} 个文件',
            f'未成功：{failed_count} 项（重命名出错 {len(error_items)}、'
            f'未匹配到文件 {len(unmatched_items)}、因匹配值重复跳过 {skipped_count}）',
        ]

        # ── 匹配值重复处理详情 ──
        if duplicate_groups:
            lines.append('')
            lines.append(f'【匹配值重复详情】列「{match_col_name}」存在重复值，'
                         '按规则仅每个重复值最先出现的一行参与重命名：')
            items = list(duplicate_groups.items())
            for i, (val, row_indices) in enumerate(items):
                if i >= 8:
                    lines.append(f'... 及其他 {len(items) - 8} 个重复值')
                    break
                lines.append(f'● 匹配值「{val}」（出现 {len(row_indices)} 次）：')
                for j, ri in enumerate(row_indices):
                    excel_row = self.roster_rows[ri].get('__excel_row__', '?')
                    o = row_outcomes.get(ri, {})
                    status = o.get('status')
                    if j == 0:
                        # 第一个重名行：报告实际结果（成功时写出原名和现名）
                        if status == 'success':
                            lines.append(f'  ✓ Excel 第 {excel_row} 行 重命名成功：'
                                         f'「{o["ori"]}」→「{o["new"]}」')
                        elif status == 'error':
                            lines.append(f'  ✗ Excel 第 {excel_row} 行 重命名失败：'
                                         f'「{o["ori"]}」— {o["error"]}')
                        else:
                            lines.append(f'  ✗ Excel 第 {excel_row} 行 未匹配到文件，未重命名')
                    else:
                        lines.append(f'  ✗ Excel 第 {excel_row} 行 未处理（与前面重名，已跳过）')

        # ── 重命名出错详情 ──
        if error_items:
            lines.append('')
            lines.append('【重命名出错】')
            for i, (ri, o) in enumerate(error_items):
                if i >= 8:
                    lines.append(f'... 及其他 {len(error_items) - 8} 个')
                    break
                excel_row = self.roster_rows[ri].get('__excel_row__', '?')
                lines.append(f'· Excel 第 {excel_row} 行：文件「{o["ori"]}」— {o["error"]}')

        # ── 未匹配到文件的行 ──
        if unmatched_items:
            lines.append('')
            lines.append('【未匹配到文件的行】')
            for i, (ri, o) in enumerate(unmatched_items):
                if i >= 8:
                    lines.append(f'... 及其他 {len(unmatched_items) - 8} 行')
                    break
                lines.append(f'· {self._describe_row(self.roster_rows[ri])}')

        if folder_note:
            lines.append('')
            lines.append(folder_note)

        text = '\n'.join(lines)
        if failed_count > 0:
            QMessageBox.warning(self, '重命名完成（存在未成功项）', text)
        else:
            QMessageBox.information(self, '重命名完成', text)


if __name__ == '__main__':
    app = QApplication([])
    ex = RenameApp()
    ex.show()
    app.exec_()
